import openpyxl
import send2trash
import os

os.chdir(".\\")

#excel object assigning
f_name="workbook.xlsx"
try:
    xls_book = openpyxl.load_workbook(f_name)
except FileNotFoundError:
    msg = "Can't find file {0}.".format(f_name)
    print(msg)

sheet_names = xls_book.get_sheet_names()
xls_sheet = xls_book.get_sheet_by_name(sheet_names[0]) #do not change names of sheets


################### WORD
from docx import Document

#delete old etiketten file
if os.path.isfile(".\\etiketten.docx"):
   send2trash.send2trash("etiketten.docx")

#load word document
document = Document("template.docx") #word doc not dynamic - template

#init variables
curr_table=0
excel_row=1

#get last excel row
while xls_sheet.cell(excel_row,1).value:
    max_excel_row=excel_row
    excel_row+=1

#make sure enough tables are available
if max_excel_row>750:
    print("Not enough tables in word document - contact support")
    msg=input("Press any key to exit")
    exit()

#set to correct table
def set_current_table(document,curr_table):
    table_obj=document.tables[curr_table]
    return table_obj

#init. adress list
adress_list=[]


#create adress list
for excel_row in range(2,max_excel_row+1): #skip header row and go until max row
    adress_col=[17,18,19,20]
    adress_text=""
    for col in adress_col: #Name and street
        if xls_sheet.cell(excel_row,col).value:
            adress_text+=str(xls_sheet.cell(excel_row,col).value)+"\n"

    adress_text+=xls_sheet.cell(excel_row,23).value + " " + xls_sheet.cell(excel_row,21).value # zip code + city

    if xls_sheet.cell(excel_row,24).value!="DE": #country
        adress_text+="\n" + xls_sheet.cell(excel_row,24).value
    if adress_text not in adress_list: # check if adress is already in adress list
        adress_list.append(adress_text)

#enter adress into word table cell
def fill_table_in_word(table_obj, adress_list,current_adress_index):
    word_columns=[0,2,4]
    count_of_adresses=len(adress_list)  # count non-duplicate adresses
    for word_row in range(9):
        for word_column in word_columns:
            if current_adress_index==count_of_adresses:
                return current_adress_index
            current_table_cell=table_obj.cell(word_row,word_column)
            current_table_cell.text=adress_list[current_adress_index]
            current_adress_index+=1
    return current_adress_index

#set to first table
table_obj=set_current_table(document,curr_table)
current_adress_index=0
count_of_adresses=len(adress_list)

#go through all deliveries
while current_adress_index<count_of_adresses:
    current_adress_index=fill_table_in_word(table_obj, adress_list,current_adress_index) #enter values for a whole table, return the newst excel row

    curr_table+=1
    table_obj=set_current_table(document,curr_table) #go to next table because prior table is full

#save document
document.save("etiketten.docx")

print('SUCCESSFULL')

msg=input("Press Enter")