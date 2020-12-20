import sys
import csv
import openpyxl
import os


os.chdir(".\\")

#read in txt file
f_name="amazon_source.txt"
try:
    with open(f_name,encoding='utf-8') as f_obj: #txt file nicht dynamisch
        tab_file = f_obj.readlines()
except FileNotFoundError:
    msg = "Can't find file {0}.".format(f_name)
    print(msg)

#external - make csv file 
reader = csv.reader(tab_file, delimiter='\t')
first_row = next(reader)
num_cols = len(first_row)
tab_reader = csv.reader(tab_file, delimiter='\t')

#excel object assigning
f_name="workbook.xlsx"
try:
    xls_book = openpyxl.load_workbook(f_name)
except FileNotFoundError:
    msg = "Can't find file {0}.".format(f_name)
    print(msg)

sheet_names = xls_book.get_sheet_names()
xls_sheet = xls_book.get_sheet_by_name(sheet_names[0]) #do not change names of sheets

###remove old data sheet
# index of [sheet_name] sheet
idx = xls_book.sheetnames.index("Sheet1")
# remove [sheet_name]
xls_book.remove_sheet(xls_sheet)
# create an empty sheet [sheet_name] using old index
xls_book.create_sheet("Sheet1", idx)

xls_sheet = xls_book.get_sheet_by_name("Sheet1") #set sheet

#external -  save csv into excel document
for row_index, row in enumerate(tab_reader):
    number = 0
    col_number = 1
    while number < num_cols:
        cell_tmp = xls_sheet.cell(row = row_index+1, column = col_number)

        cell_tmp.value = row[number]
        number += 1
        col_number += 1

xls_book.save("workbook.xlsx") #excel file not dynamic

xls_book.close()
f_obj.close()

print('SUCCESSFULL')


msg=input("Press Enter")