import sys
import csv
import openpyxl

#read in txt file
f_name="lena.txt"
try:
    with open(f_name) as f_obj: #txt file nicht dynamisch
        tab_file = f_obj.readlines()
except FileNotFoundError:
    msg = "Can't find file {0}.".format(f_name)
    print(msg)

#external - make csv file
reader = csv.reader(tab_file, delimiter='\t')
first_row = next(reader)
num_cols = len(first_row)
tab_reader = csv.reader(tab_file, delimiter='\t')

#excel object assigning
f_name="workbook.xlsx"
try:
    xls_book = openpyxl.load_workbook(f_name)
except FileNotFoundError:
    msg = "Can't find file {0}.".format(f_name)
    print(msg)

sheet_names = xls_book.get_sheet_names()
xls_sheet = xls_book.get_sheet_by_name(sheet_names[0]) #do not change names of sheets

###remove old data sheet
# index of [sheet_name] sheet
idx = xls_book.sheetnames.index("Sheet1")
# remove [sheet_name]
xls_book.remove_sheet(xls_sheet)
# create an empty sheet [sheet_name] using old index
xls_book.create_sheet("Sheet1", idx)

xls_sheet = xls_book.get_sheet_by_name("Sheet1") #set sheet

#external -  save csv into excel document
for row_index, row in enumerate(tab_reader):
    number = 0
    col_number = 1
    while number < num_cols:
        cell_tmp = xls_sheet.cell(row = row_index+1, column = col_number)
        cell_tmp.value = row[number]
        number += 1
        col_number += 1

xls_book.save("workbook.xlsx") #excel file not dynamic

################### WORD

from docx import Document

#load word document
document = Document("template.docx") #word doc not dynamic - template

#init variables
curr_table=0
excel_row=1

#get last excel row
while xls_sheet.cell(excel_row,1).value:
    max_excel_row=excel_row
    excel_row+=1

#skip header excel row
excel_row=2

#make sure enough tables are available
if max_excel_row>750:
    print("Not enough tables in word document - contact support")
    exit()

#set to correct table
def set_current_table(document,curr_table):
    table_obj=document.tables[curr_table]
    return table_obj

#connect the adress
def connect_text(excel_row):
    adress_col=[17,18,19,20,23,21,24]
    adress_text=""
    for col in adress_col:
        if xls_sheet.cell(excel_row,col).value:
            adress_text+=str(xls_sheet.cell(excel_row,col).value)+"\n"
    return adress_text
    
#enter adress into word table cell
def fill_table_in_word(table_obj, excel_row):
    word_cols=[0,2,4]
    for word_row in range(9):
        for word_col in word_cols:
            adress_string=connect_text(excel_row)
            curr_cell=table_obj.cell(word_row,word_col)
            curr_cell.text=adress_string
            excel_row+=1
    return excel_row

#set to first table
table_obj=set_current_table(document,curr_table)

#go through all deliveries
while (excel_row<=max_excel_row):
    excel_row=fill_table_in_word(table_obj, excel_row) #enter values for a whole table, return the newst excel row

    curr_table+=1
    table_obj=set_current_table(document,curr_table) #go to next table because prior table is full

#save document
document.save("etiketten.docx")

xls_book.close()
f_obj.close()

print('SUCCESSFULL (most likely)')
print("Das macht dann 12.000 Mark")

msg=input("Press Enter")

#bestellungen sind nicht zusammengefasst
#max 750 Bestellungen
#copy table - create table?