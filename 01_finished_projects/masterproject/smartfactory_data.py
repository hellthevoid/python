import csv
import openpyxl
import os
import sys

def find_last_row(wb,sheet_num,start_row,column):
    sheet_names = wb.sheetnames
    data_sheet=wb[sheet_names[sheet_num]]

    for row_index, row in enumerate(data_sheet.iter_rows(min_row=start_row,max_row=1000000,min_col=column,max_col=column)):
        for cell in row:
            if cell.value is None:
                return row_index
    print("Error, function failed")  
    exit()
    return 0

#function to find last column of a row
def find_last_col(wb,sheet_num,start_col,row):
    sheet_names = wb.sheetnames
    data_sheet=wb[sheet_names[sheet_num]]

    for col_index, col in enumerate(data_sheet.iter_cols(min_col=start_col,max_col=10000,min_row=row,max_row=row)):
        for cell in col:
            if cell.value is None:
                return col_index
    print("Error, function failed")  
    exit()
    return 0

if __name__ == '__main__': #only run this code below if the script is run directly and not used as a module

    print("Calculation is starting...")
    #set folder of script to current working directory
    file_path=os.path.realpath(__file__)
    directory_path=os.path.dirname(file_path)
    os.chdir(directory_path) 

    #insert path of module to sys.path to access it from other modules
    if directory_path not in sys.path:
        sys.path.insert(0, directory_path) 

    #open the excel template
    f_name="smartfactory_template.xlsx"
    try:
        xls_book = openpyxl.load_workbook(f_name)
    except FileNotFoundError:
        msg = "Can't find file {0}.".format(f_name)
        print(msg)

    #set sheet objects
    sheet_names = xls_book.sheetnames

    data_sheet = xls_book[sheet_names[0]]
    analysis_sheet=xls_book[sheet_names[1]]
    dld_sheet=xls_book[sheet_names[2]]

    #set path for data csv files
    path=".\\data"

    #reading all csv files in a folder and put them into excel file (+trimming)
    curr_col=1
    for folderName,subfolders,fNames in os.walk(path):
        
        for fname in fNames:
            with open(os.path.join(folderName,fname)) as f_obj:
                reader = csv.reader(f_obj, delimiter=',')
                for row_index, row in enumerate(reader):
                    for col_index, col in enumerate(row):
                        if col_index!=0:
                            cell=data_sheet.cell(row_index+1,col_index+1+curr_col)
                            cell.value=col.strip() #remove start/end spaces
            curr_col+=2

    #delete two unused columns
    data_sheet.delete_cols(2)
    data_sheet.delete_cols(3)

    #enter time in first column
    time=0
    new_max_row=find_last_row(xls_book,0,2,3) #TO DO: fix find lastrow
    data_sheet["A1"].value="Time"

    #enters time into first column
    for row in data_sheet.iter_rows(min_row=2,max_row=new_max_row+1,max_col=1):
        for cell in row:
            cell.value=time
            time+=0.1

    #ANALYSIS
    first_cell=0
    #go through each important sensor column 
    for col_num in range(4,13,2): #using the range function with the step method -> important_cols=[4,6,8,10,12]

        previous_cell_value=""

        analysis_row_index=2 #row in analysis sheet
        
        #go through each row of a column by using .iter_rows
        for row_index, row in enumerate(data_sheet.iter_rows(min_row=2,max_row=new_max_row,min_col=col_num,max_col=col_num)):
            for cell in row:
                #check previous value is false and current one is true -> start of a package on a  working station
                if cell.value=="true" and previous_cell_value=="false": #algorithm
                        ##assign time value (column 1) to cell in new sheet (analysis_sheet)
                        analysis_cell=analysis_sheet.cell(analysis_row_index,col_num)

                        analysis_cell.value=data_sheet.cell(row_index+2,1).value

                        #get the first value and substract it from all the other values to null the measurement
                        if first_cell==0: #indicator to find the first written cell
                            first_value=data_sheet.cell(row_index+2,1).value
                            first_cell+=1
                        analysis_cell.value=analysis_cell.value-first_value

                        print(analysis_cell.value)
                        analysis_row_index+=1

                previous_cell_value=cell.value #save cell value for checking on next row

    analysis_sheet.delete_cols(1,amount=3)

    #write the headers for the analysis sheet
    index=1
    for col_index,col in enumerate(analysis_sheet.iter_cols(min_row=2,max_row=2,max_col=30)):
        for cell in col:
            if cell.value is not None:
                analysis_sheet.cell(1,col_index+1).value="AS_{} ZU".format(index)
                index+=1
                
    for col_index,col in enumerate(analysis_sheet.iter_cols(min_row=2,max_row=2,max_col=30)):
        for cell in col:
            if cell.value is None:
                analysis_sheet.delete_cols(col_index+1)

    last_col=find_last_col(xls_book,1,1,1)

    #enter the last processing time to calculate the departure of packages
    processing_time_of_last_station=float(input("Enter the processing time of the last work station please!"))

    #write header for Abgang
    analysis_sheet.cell(1,last_col+1).value="AS_{0} Abgang".format(index-1)

    #find the last row of the last station
    last_row=find_last_row(xls_book,1,1,last_col)

    #write a new column with the departure times
    for row_index,row in enumerate(analysis_sheet.iter_rows(min_row=2,max_row=last_row,min_col=last_col+1,max_col=last_col+1)):
        for cell in row:
            cell.value=float(analysis_sheet.cell(row_index+2,last_col).value or 0)
            cell.value=cell.value+processing_time_of_last_station

    #save and close workbook
    xls_book.save("smartfactory_data.xlsx") #excel file not dynamic
    xls_book.close()

    print('SUCCESSFULL')

    msg=input("Press Enter")