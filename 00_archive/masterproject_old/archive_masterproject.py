import sys
import csv
import openpyxl
import send2trash
import os
import shutil

os.chdir(".\\")

# #read in txt file
# f_name="amazon_source.txt"
# try:
#     with open(f_name) as f_obj: #txt file nicht dynamisch
#         tab_file = f_obj.readlines()
# except FileNotFoundError:
#     msg = "Can't find file {0}.".format(f_name)
#     print(msg)

# #external - make csv file
# reader = csv.reader(tab_file, delimiter='\t')
# first_row = next(reader)
# num_cols = len(first_row)
# tab_reader = csv.reader(tab_file, delimiter='\t')

#excel object assigning

def find_last_row(wb,sheet_num,start_row,column):
    xls_sheet=wb.get_sheet_by_name(sheet_names[sheet_num])

    row=start_row
    while xls_sheet.cell(row,column)!="":
        row+=1
    return row

f_name="masterproject.xlsx"
try:
    xls_book = openpyxl.load_workbook(f_name)
except FileNotFoundError:
    msg = "Can't find file {0}.".format(f_name)
    print(msg)

sheet_names = xls_book.get_sheet_names()
xls_sheet = xls_book.get_sheet_by_name(sheet_names[0])#do not change names of sheets

###remove old data sheet
# index of [sheet_name] sheet
idx = xls_book.sheetnames.index("Sheet1")
# remove [sheet_name]
xls_book.remove_sheet(xls_sheet)
# create an empty sheet [sheet_name] using old index
xls_book.create_sheet("Sheet1", idx)

xls_sheet = xls_book.get_sheet_by_name("Sheet1") #set sheet

path="C:\\Users\\sonyx\\Desktop\\python\\masterproject\\data"

col_number=2

for folderName,subfolders,fNames in os.walk(path):
    for fname in fNames:
        try:
            with open(os.path.join(folderName,fname)) as f_obj: #txt file nicht dynamisch
                tab_file = f_obj.readlines()
        except FileNotFoundError:
            msg = "Can't find file {0}.".format(fname)
            print(msg)

        reader = csv.reader(tab_file, delimiter=';')
        first_row = next(reader)
        num_cols = len(first_row)
        tab_reader = csv.reader(tab_file, delimiter=';')


        for row in xls_sheet.iter_rows(min_row=1,max_col=3,max_row=3):
                for cell in row:
                    pass
                    
        for row_index, row in enumerate(tab_reader):
            number = 0
            if row_index!=0:
                col_number-=num_cols
            while number < num_cols:
                cell_tmp = xls_sheet.cell(row = row_index+1, column = col_number)
                cell_tmp.value = row[number]
                number += 1
                col_number += 1
        col_number+=2

last_row=find_last_row(xls_book,0,2,1)

time=0
for row in range(last_row):
    xls_sheet.cell(row,0).value=time
    time+=0.1


# ###remove old data sheet
# # index of [sheet_name] sheet
# idx = xls_book.sheetnames.index("Sheet1")
# # remove [sheet_name]
# xls_book.remove_sheet(xls_sheet)
# # create an empty sheet [sheet_name] using old index
# xls_book.create_sheet("Sheet1", idx)

# xls_sheet = xls_book.get_sheet_by_name("Sheet1") #set sheet

#external -  save csv into excel document


xls_book.save("masterproject.xlsx") #excel file not dynamic

xls_book.close()


print('SUCCESSFULL (most likely)')

msg=input("Press Enter")