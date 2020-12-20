#function to find last row of a column
import openpyxl

def find_last_row(wb,sheet_num,start_row,column):
    sheet_names = wb.sheetnames
    data_sheet=wb[sheet_names[sheet_num]]

    for row_index, row in enumerate(data_sheet.iter_rows(min_row=start_row,max_row=1000000,min_col=column,max_col=column)):
        for cell in row:
            if cell.value is None:
                return row_index
    print("Error, function failed")  
    exit()
    return 0

#function to find last column of a row
def find_last_col(wb,sheet_num,start_col,row):
    sheet_names = wb.sheetnames
    data_sheet=wb[sheet_names[sheet_num]]

    for col_index, col in enumerate(data_sheet.iter_cols(min_col=start_col,max_col=10000,min_row=row,max_row=row)):
        for cell in col:
            if cell.value is None:
                return col_index
    print("Error, function failed")  
    exit()
    return 0