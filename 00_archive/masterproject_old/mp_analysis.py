import openpyxl
import os
import sys


#set folder of script to current working directory
file_path=os.path.realpath(__file__)
directory_path=os.path.dirname(file_path)
os.chdir(directory_path)

print (sys.path)
if directory_path not in sys.path:
    sys.path.insert(0, directory_path) #insert path of module to sys.path to access it from other modules

#TO DO:
# add path to sys.path ->dynamic    

#custom modules path must be included into PATH or PYTHONPATH of the system (do so by using sys.path)
from mp_data import find_last_row, find_last_col

print(os.getcwd())

#open the excel book
f_name="masterproject_data.xlsx"
try:
    xls_book = openpyxl.load_workbook(f_name)
except FileNotFoundError:
    msg = "Can't find file {0}.".format(f_name)
    print(msg)

#assigning the sheet and object name
sheet_names = xls_book.get_sheet_names()
data_sheet = xls_book.get_sheet_by_name(sheet_names[0])
analysis_sheet=xls_book.get_sheet_by_name(sheet_names[1])
dld_sheet=xls_book.get_sheet_by_name(sheet_names[2])

#AS1 -ignore first true section- start with second true section -> ring 6 (delte first list entry?!)

print("Transfering is getting started")

#Algorithm
#is it possible to load only values??
#what about false alarms? (not important in first scenario)
#filled or unfilled system?

new_max_row=find_last_row(xls_book,0,2,3)

#go through each column
for col_num in range(4,13,2): #using the range function with the step method -> important_cols=[4,6,8,10,12]

    previous_cell_value=""
    analysis_row_index=2 #row in analysis sheet

    #go through each row of a column by using .iter_rows - min/max row are static!
    for row_index, row in enumerate(data_sheet.iter_rows(min_row=2,max_row=new_max_row,min_col=col_num,max_col=col_num)):
        for cell in row:
            
            #check previous value is false and current one is true -> start of a package on a  working station
            if cell.value=="true" and previous_cell_value=="false": #algorithm
                    ##assign time value (column 1) to cell in new sheet (analysis_sheet)
                    analysis_sheet.cell(analysis_row_index,col_num).value=data_sheet.cell(row_index+2,1).value 
                    
                    print(data_sheet.cell(row_index+2,1).value)
                    analysis_row_index+=1

            previous_cell_value=cell.value #save cell value for checking on next row
    print("transfering for AS is done")

analysis_sheet.delete_cols(1,amount=3)
index=1
for col_index,col in enumerate(analysis_sheet.iter_cols(min_row=2,max_row=2,max_col=30)):
    for cell in col:
        if cell.value is not None:
            analysis_sheet.cell(1,col_index+1).value="AS_{} ZU".format(index)
            index+=1
            
for col_index,col in enumerate(analysis_sheet.iter_cols(min_row=2,max_row=2,max_col=30)):
    for cell in col:
        if cell.value is None:
            analysis_sheet.delete_cols(col_index+1)

last_col=find_last_col(xls_book,1,1,1)

abgang=float(input("Enter the processing time of the last work station please!"))

#write header for Abgang
analysis_sheet.cell(1,last_col+1).value="AS_{0} Abgang".format(index-1)

last_row=find_last_row(xls_book,1,1,last_col)

for row_index,row in enumerate(analysis_sheet.iter_rows(min_row=2,max_row=last_row,min_col=last_col+1,max_col=last_col+1)):
    for cell in row:

        cell.value=float(analysis_sheet.cell(row_index+2,last_col).value or 0)
        cell.value=cell.value+abgang


#save excel file
xls_book.save("masterproject_data.xlsx")
print("Analysis succesful")

msg=input("Enter any key to exit")

xls_book.close()
