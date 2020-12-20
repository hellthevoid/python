import csv
import openpyxl
import send2trash
import os

#function to find last row of soemthing
def find_last_row(wb,sheet_num,start_row,column):
    sheet_names = wb.get_sheet_names()
    xls_sheet=wb.get_sheet_by_name(sheet_names[sheet_num])

    for row_index, row in enumerate(xls_sheet.iter_rows(min_row=start_row,min_col=column,max_col=column)):
        for cell in row:
            if cell.value==None:
                return row_index+1
    print("Error, function failed")  
    exit()
    return 0

if __name__ == '__main__': #only run this code below if the script is run directly


    #set folder of script to current working directory
    file_path=os.path.realpath(__file__)
    directory_path=os.path.dirname(file_path)
    os.chdir(directory_path) 

    #open a workbook
    f_name="masterproject_template.xlsx"
    try:
        xls_book = openpyxl.load_workbook(f_name)
    except FileNotFoundError:
        msg = "Can't find file {0}.".format(f_name)
        print(msg)
        msg=input("Enter any key to exit")
        exit()

    #get sheet object
    sheet_names = xls_book.get_sheet_names()
    xls_sheet = xls_book.get_sheet_by_name(sheet_names[0])#do not change names of sheets

    path=".\\data"

    #reading all csv files ina  folder  and put them into excel file (+trimming)
    curr_col=1

    for folderName,subfolders,fNames in os.walk(path):
        for fname in fNames:
            with open(os.path.join(folderName,fname)) as f_obj: #txt file nicht dynamisch
                reader = csv.reader(f_obj, delimiter=',')
                for row_index, row in enumerate(reader):
                    for col_index, col in enumerate(row):
                        if col_index!=0:
                            cell=xls_sheet.cell(row_index+1,col_index+1+curr_col)
                            cell.value="=TRIM(\"{0}\")".format(col)
            curr_col+=2

    #delete two unused columns
    xls_sheet.delete_cols(2)
    xls_sheet.delete_cols(3)

    #enter time in first column
    time=0
    new_max_row=find_last_row(xls_book,0,2,3)

    for row in xls_sheet.iter_rows(min_row=2,max_row=new_max_row,max_col=1):
        for cell in row:
            cell.value=time
            time+=0.1

    #enter header into first row

    """
    header=["time","start_sensor","sensor_wait_AS1","sensor_AS1","sensor_wait_AS2","sensor_AS2","sensor_wait_AS3","sensor_AS3","sensor_wait_AS4","sensor_AS4","sensor_wait_AS5","sensor_AS5"]
    i=0
    for col in xls_sheet.iter_cols(max_row=1,max_col=12):
        for cell in col:
            cell.value=header[i]
            i+=1
    """
    #save and close workbook
    xls_book.save("masterproject_data.xlsx") #excel file not dynamic
    xls_book.close()

    print('SUCCESSFULL (most likely)')

    msg=input("Press Enter")